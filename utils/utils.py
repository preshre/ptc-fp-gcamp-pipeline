"""
Utility functions for PTC fiber photometry analysis.
"""

import re
import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment


def resolve_long_path(path):
    """
    Resolve long paths for Windows systems.
    
    Parameters:
    -----------
    path : str
        File path
        
    Returns:
    --------
    Path : Resolved path object
    """
    if os.name == 'nt':
        return Path(f"\\\\?\\{path}")
    return Path(path)


def parse_text_file_auc(file_path):
    """
    Parse AUC log file and organize data by experiment type.
    
    Expected format: MouseName-Experiment: [[Pre-CS values], [CS values], [Post-CS values]]
    
    Parameters:
    -----------
    file_path : str
        Path to auc.txt file
        
    Returns:
    --------
    dict : Data organized by experiment type (Training, LTM1d, LTM14d, LTM28d)
    """
    data = {}
    num_cs_per_experiment = {}
    
    with open(file_path, 'r') as file:
        for raw_line in file:
            line = raw_line.strip()
            if not line:
                continue
            
            header, rest = line.split(': ', 1)
            mouse_name, experiment = header.rsplit('-', 1)
            
            # Extract arrays from brackets
            array_strings = re.findall(r'\[([-\d., ]+)\]', rest)
            arrays = [list(map(float, arr_str.split(', '))) for arr_str in array_strings]
            
            if len(arrays) != 3:
                raise ValueError(f"Expected 3 arrays (Pre, CS, Post), got {len(arrays)}")
            
            pre_vals, cs_vals, post_vals = arrays
            
            if not (len(pre_vals) == len(cs_vals) == len(post_vals)):
                raise ValueError(f"Mismatched array lengths in line: {line}")
            
            num_cs = len(pre_vals)
            
            # Initialize experiment structure
            if experiment not in data:
                num_cs_per_experiment[experiment] = num_cs
                data[experiment] = {'Mouse Name': []}
                for cs_idx in range(1, num_cs + 1):
                    data[experiment][f'Pre-CS {cs_idx}'] = []
                    data[experiment][f'CS {cs_idx}'] = []
                    data[experiment][f'Post-CS {cs_idx}'] = []
            
            # Add data
            data[experiment]['Mouse Name'].append(mouse_name)
            for cs_idx in range(num_cs):
                data[experiment][f'Pre-CS {cs_idx + 1}'].append(pre_vals[cs_idx])
                data[experiment][f'CS {cs_idx + 1}'].append(cs_vals[cs_idx])
                data[experiment][f'Post-CS {cs_idx + 1}'].append(post_vals[cs_idx])
    
    return data


def convert_to_excel_with_formatting_auc(data, output_file):
    """
    Write AUC data to formatted Excel file.
    
    Parameters:
    -----------
    data : dict
        Data from parse_text_file_auc
    output_file : str
        Output Excel file path
    """
    # Write raw data
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for experiment, df_data in data.items():
            df = pd.DataFrame(df_data)
            df.to_excel(writer, sheet_name=experiment, index=False)
    
    # Format workbook
    workbook = load_workbook(output_file)
    metrics_per_cs = 3  # Pre-CS, CS, Post-CS
    sub_labels = ["Pre-CS", "CS", "Post-CS"]
    
    for sheet_name in data.keys():
        sheet = workbook[sheet_name]
        total_data_cols = sheet.max_column - 1
        num_cs = total_data_cols // metrics_per_cs
        
        # Insert sub-header row
        sheet.insert_rows(2)
        
        for cs_idx in range(1, num_cs + 1):
            start_col = 2 + (cs_idx - 1) * metrics_per_cs
            end_col = start_col + metrics_per_cs - 1
            
            # Merge CS header
            sheet.merge_cells(start_row=1, start_column=start_col, 
                            end_row=1, end_column=end_col)
            cell = sheet.cell(row=1, column=start_col)
            cell.value = f"CS{cs_idx}"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Add sub-labels
            for offset, label in enumerate(sub_labels):
                c = sheet.cell(row=2, column=start_col + offset)
                c.value = label
                c.alignment = Alignment(horizontal="center", vertical="center")
        
        # Center-align data
        for row in range(3, sheet.max_row + 1):
            for col_idx in range(2, 2 + num_cs * metrics_per_cs):
                sheet.cell(row=row, column=col_idx).alignment = Alignment(
                    horizontal="center", vertical="center"
                )
    
    workbook.save(output_file)


def parse_text_file_ymax(file_path):
    """
    Parse Ymax log file and organize data by experiment type.
    
    Expected format: MouseName-Experiment: value1, value2, ...
    
    Parameters:
    -----------
    file_path : str
        Path to yMax.txt file
        
    Returns:
    --------
    dict : Data organized by experiment type
    """
    data = {}
    
    with open(file_path, 'r') as file:
        for raw_line in file:
            line = raw_line.strip()
            if not line:
                continue
            
            mouse_info, values_str = line.split(': ', 1)
            mouse_name, experiment_type = mouse_info.split('-')
            
            values = [float(x) for x in re.findall(r'[-+]?\d*\.\d+|\d+', values_str)]
            
            if experiment_type not in data:
                data[experiment_type] = {'Mouse Name': []}
                if experiment_type == 'Training':
                    data[experiment_type].update({
                        'CS1': [], 'CS2': [], 'US1': [], 'US2': []
                    })
                else:
                    data[experiment_type].update({
                        'CS1': [], 'CS2': [], 'CS3': []
                    })
            
            data[experiment_type]['Mouse Name'].append(mouse_name)
            
            if experiment_type == 'Training':
                data[experiment_type]['CS1'].append(values[0])
                data[experiment_type]['CS2'].append(values[1])
                data[experiment_type]['US1'].append(values[2])
                data[experiment_type]['US2'].append(values[3])
            else:
                data[experiment_type]['CS1'].append(values[0])
                data[experiment_type]['CS2'].append(values[1])
                data[experiment_type]['CS3'].append(values[2])
    
    return data


def convert_to_excel_with_formatting_ymax(data, output_file):
    """
    Write Ymax data to formatted Excel file.
    
    Parameters:
    -----------
    data : dict
        Data from parse_text_file_ymax
    output_file : str
        Output Excel file path
    """
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for experiment_type, exp_data in data.items():
            df = pd.DataFrame(exp_data)
            df.to_excel(writer, sheet_name=experiment_type, index=False)
    
    workbook = load_workbook(output_file)
    
    for experiment_type in data.keys():
        sheet = workbook[experiment_type]
        
        # Delete non-Mouse Name columns and rebuild
        for col in range(sheet.max_column, 1, -1):
            if sheet.cell(row=1, column=col).value != 'Mouse Name':
                sheet.delete_cols(col)
        
        if experiment_type == 'Training':
            columns = ['CS1', 'CS2', 'US1', 'US2']
        else:
            columns = ['CS1', 'CS2', 'CS3']
        
        # Add columns
        current_col = 2
        for col_name in columns:
            sheet.cell(row=1, column=current_col).value = col_name
            sheet.cell(row=1, column=current_col).alignment = Alignment(horizontal="center")
            current_col += 1
        
        # Add data
        row = 2
        for mouse_idx in range(len(data[experiment_type]['Mouse Name'])):
            current_col = 2
            for col_name in columns:
                sheet.cell(row=row, column=current_col).value = data[experiment_type][col_name][mouse_idx]
                sheet.cell(row=row, column=current_col).alignment = Alignment(horizontal="center")
                current_col += 1
            row += 1
    
    workbook.save(output_file)
